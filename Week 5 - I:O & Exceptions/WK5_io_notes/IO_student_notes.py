
##################################
####### Week 5. File I/O #########
##################################

'''
File I/O (input/output) is for reading information from disc (hard/ss drive, persistant) into Python. Data is stored in local memory
(on system, temporary) once it's opened into Python environment. 

Learning Outcomes - you should be able to:

- open text and csv files from a local directory on your computer
- distinguish between read, write, append methods to support your analyses
- process file information as needed
'''



###############################
#------ Syntax for Open() ----#
###############################

'''
Basic Open() Syntax

open() is a Python built-in function - https://docs.python.org/3/library/functions.html:

file = open(filepath.filename, mode) #--> where mode is read ('r'), write ('w') or append ('a')

for line in file:  # we can treat file as an iterable object
    print(line)

file.close()
'''

##############################
#----- Reading Txt Files ----#
##############################

# Example 1.  Open a Text File in read mode and print each Line. The following file is located the a working directory.

file = open('moby_dick.txt', 'r')
for line in file: 
    print(line, end = ' ') # remove end = ' ' and compare output, replaces newline character with a single space, the next print will output on the same line immediately following the space
file.close() 

'''
Issues with using open() 

1.  If you open a lot of files without closing them, you might consume too much system memory, which can slow down or even crash your program

2. Some changes to files are not always immediately written to disk; they are kept in a buffer and written later for efficiency. 
    If you don't close a file, some changes might never be written to disk, which can lead to data loss or corruption.

3. Some operating systems prevent certain operations on open files. For example, you might not be able to delete or move an open file.

'''

###########################################################
#----------- Context Manager 'with open()' ---------------#
###########################################################

# #xample 2. "with open" context manager

with open('demo.txt', 'r', encoding="utf-8") as file: 
	for line in file:  
		print(line, end=' ')


#####################################################
####------ read(), readline(), readlines() ------####
#####################################################

# # Example 1. read() = reads entire txt file

with open('moby_dick.txt', 'r', encoding="utf-8") as file: 
	moby_dick = file.read() 

print(moby_dick, '\n', f'The length of read(moby_dick) is {len(moby_dick)} and the type is {type(moby_dick)}')  


with open('moby_dick.txt', 'r', encoding="utf-8") as file: 
	content = file.read(150) # what does passing in 150 do here? Check your print-out.
	print(content, end='')

###############################################################################

# Example 3. methods that can be useful for navigating a read file

with open('moby_dick.txt', 'r') as file:
    file_content = file.read(300) # read first 300 chars
    print(file.tell()) # gives current position, gives 301 so one of the first 300 chars is a multi-byte char (> 1 byte memory; e.g., \n). 
    print(file.seek(0)) # moves position to beginning of file
    print(file.tell())
    
    file_content = file.read(50) # read next 50 chars
    print(file_content)

    file.seek(22) # moves file position to the 23rd byte (positions are 0 indexed)
    file_content = file.read(10) # reads next to chars from current, 23rd byte
    print(file_content)
    print(file.tell())

##########################################################

# Example 4. readline() --> read individual lines one-by-one

'''
Try running the with open() statement below with different print statements commented out.
Work through the prints sequentially
'''

with open('demo.txt', 'r', encoding="utf-8") as file:
	#print(file.readline())
	print(file.readline(4)) # char level, this will produce the same file.readline()[0:4]
	print(file.readline()*3)
	for line in range(2): # from current position
		print(file.readline(), end = '') # try without end = ''

with open('demo.txt', 'r', encoding="utf-8") as file: # use range() rather than the following
	print(file.readline())
	print(file.readline())

############################################################

# # Example 4.b. using implicit 'readline' in a loop with input(), allows user to step through a text 

with open('demo.txt', 'r') as file:
    for line in file:
        print(line, end=' ') #ends line with empty string, no new line character
        input() # pauses the program until the user presses enter in the terminal, ctrl+c to stop program

########################################################
# Example 5.  readlines() --> reads file and returns a list of strings

'''
Try running with different print statements commented out.
Work through the prints sequentially
'''

with open('short_dictionary.txt', 'r', encoding="utf-8") as file:
	print(file.readlines())
	
    file.seek(0) # absent this the next line returns a [] because position is end of file
	#print(file.readlines()[0:10])
	file.seek(0)
	demo = file.readlines() # with initial readlines, position is end of file and we get an empty list.
	# can use seek to get around this: file.seek(0)
	print(demo, '\n\n', type(demo))

#########################################################

# Example 6. Cleaning readlines() print output

''' 
To run line in demo, demo needs to be instantiated in Example 5
I've included some string methods here (rstrip, replace) - we'll dive into these methods when we deep-dive into Python strings

'''

for line in demo: # iterate over txt object --> note spacing in print out
	print(line)
	
# Now remove newlines in print out

for line in demo:
	print(line, end='') # prints each line separately due to for loop, '' = end with empty string rather than newline, even though print isnt adding newlines, text already has newlines in it
	print(line.replace('\n', '')) # we remove newlines, but print adds them

for line in demo: # remove newline before printing
    print(line.rstrip('\n'), end='')

##############################
#----- Writing Txt Files ----#
##############################

''' 
With write mode, create a Text File and add text. 
'''

# Example 7. write() method -- create a file and write in it. 

'''The following writes 'write_demo.txt' to the current working directory'''

with open('write_demo.txt', 'w', encoding="utf-8") as writer: 
	writer.write('This is a new text file with created in write mode!\n')  # .write() -> writes single string
	writer.write('This is an additional line of text.\n')
	writer.write('This is the final line of text!\n')
	writer.write('Well, not really!')


''' 
BE CAREFUL - .write() will over-write information in existing files.

Substitute 'x' for 'w' as overwrite safeguard, will throw error if file exists
'''

############################################################

# Example 8. Appending information to a file without overwriting

'''
Use .append() connection to avoid over-writing an existing file

.append() will create new file if file doesn't exist.
'''

with open('write_demo.txt', 'a', encoding="utf-8") as appended: 
	appended.write('\n\nThis line has been appended to write_demo.txt.') 


# Example 9 Open a text file and write it to another directory one level up

with open('demo.txt', 'r', encoding='utf-8') as file:
	text = file.read()

with open('../demo_copy.txt', 'w', encoding = 'utf-8') as file:  
	file.write(text)
    
# Example 9b. A more Pythonic way to read/write simultaneously

input_file = Path("input_txt")
output_file = Path("output/output_txt") # to write to parent dir "../output/output_txt", to write to subdirectory "subdir_name/subdir2_name/filename"

output_file.parent.mkdir(parents=True, exist_ok=True) # parents, exists

with open(input_file, 'r') as infile, open(output_file, 'w') as outfile:
    # allows for line by line processing
    for line in infile:
        # can create alterations here if we want to. 
        outfile.write(line)

	
# Example 10. Read and Write Sequentially -- we didn't cover this but it's just an extension of what we've done
	
with open('demo.txt', 'r', encoding="utf-8") as demo_read:
	line_lst = [] # empty list that we populate during for loop
	index = 1
	for line in demo_read:
		add_index = str(index)+'.' + ' ' + line + '\n'  # add line numbering
		line_lst.append(add_index)
		index += 1

with open('write_demo.txt', 'w', encoding="utf-8") as demo_write:
	demo_write.writelines(line_lst) # writelines writes a list of strings to a opened file in order

# See also:  https://stackoverflow.com/questions/9282967/how-to-open-a-file-using-the-open-with-statement

##############################
#--------- CSV Files --------#
##############################

# Example 11. Read and format a csv file - essentially a comma delimited text file

with open('demo_data.csv', 'r', encoding = 'utf-8') as file:
	data = []
	for line in file.readlines():
		data.append(line)
print(data)


# Compare to the next examples. What is the difference?


with open('demo_data.csv', r) as file:
    for line in file:
        print(line, end = '')
        
# OR

with open('demo_data.csv', 'r') as file:
	data = []
	for line in file.readlines():
		data.append(line.lower().replace('\n', '').split(','))
print(data)



# Example 12. With list comprehension - - we'll cover this in an upcoming class, take a look at the code and see if you can interpret the syntax.

data = [i.lower().replace('\n', '').split(',') for i in file.readlines()]  # clean text, creates list of lists
print(data)

########################################################## 

# Example 13. We can loop through the data if it has been saved to a reference name

for row in data:
	print(row)

###########################################################

# Example 14. print column names -- our first row in the dataset

columns = data[0]
print(f'These are the column names: {columns}')

############################################################

# Example 15. This includes an example of list slicing. This will be covered in detail when we take a deeper dive into lists.  Select 'age' column and save into a list


age = []
for vars in data [1:]: # all rows from 1 to end, 0 -> actual row 1
	age.append(vars[0]) # list method .append()
print(age)

# Try doing the same for hometown and name

#############################################################

# Example 16. Calculate the average age -- we didn't cover this but it's a good example for study

age = []
for vars in data [1:]: # what is this slice doing?
	age.append(int(vars[0])) # what is going on here

print(age)
print(len(age))
print(f'The average age is: {sum(age)/len(age)}') 

with open('data_age.csv', 'w') as file:  # saving as csv
 	file.write(str(age))

with open('data_age.csv', 'a') as file:
    #file.seek(0) # what happens if we uncomment and this as well?
    file.write('\n\n README: this is the age data from demo_data.csv')

    
    
########################################################## 

# CSV Package - The csv module’s reader and writer objects read and write sequences. Programmers can also read and write data in dictionary form using the DictReader and DictWriter classes



########################################################## 

# CSV Package - The csv module’s reader and writer objects read and write sequences. 
# Programmers can also read and write data in dictionary form using the DictReader and DictWriter classes
# Each row read from the csv file is returned as a list of strings. No automatic data type conversion is 
# performed unless the QUOTE_NONNUMERIC format option is specified (in which case unquoted fields are 
# transformed into floats).

# alwyas use newline = '' when opening files. 

import csv

def read_csv_file(file):
    with open(file, 'r', newline = '', encoding = 'utf-8') as f:
        f = csv.reader(f)
        header = next(f)
        print(f'Header: {header}')
        
        # create list of str from dataset
        data = []
        for row in f:
            data.append(row)
            print(row)

read_csv_file('demo_data.csv')


# Writing a CSV file
def write_csv_file(filename, header, data):
    """Write data to a CSV file"""
    with open(filename, 'w', newline='', encoding='utf-8') as f:
        csv_writer = csv.writer(f)
        
        # Write header
        csv_writer.writerow(header)
        
        # Write data rows
        csv_writer.writerows(data)
        print(f"Data written to {filename}")
        
     
##########################################################    

# openpyxl - this package allows us to work with Excel files
# The follow were coded with Claude

import csv
from openpyxl import Workbook, load_workbook

# Reading an Excel file
def read_excel_file(filename):
    """Read data from an Excel file using openpyxl"""
    workbook = load_workbook(filename)
    sheet = workbook.active  # Get the active sheet
    
    print(f"Reading from sheet: {sheet.title}")
    
    data = []
    for row in sheet.iter_rows(values_only=True):
        print(row)
        data.append(row)
    
    workbook.close()
    return data

# Writing an Excel file
def write_excel_file(filename, header, data):
    """Write data to an Excel file using openpyxl"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    
    # Write header
    sheet.append(header)
    
    # Write data rows
    for row in data:
        sheet.append(row)
    
    workbook.save(filename)
    print(f"Data written to {filename}")

# Converting between CSV and Excel
def csv_to_excel(csv_filename, excel_filename):
    """Convert CSV file to Excel"""
    # Read CSV
    with open(csv_filename, 'r', newline='', encoding='utf-8') as csv_file:
        csv_reader = csv.reader(csv_file)
        
        # Create Excel workbook
        workbook = Workbook()
        sheet = workbook.active
        
        # Write CSV data to Excel
        for row in csv_reader:
            sheet.append(row)
        
        workbook.save(excel_filename)
        print(f"Converted {csv_filename} to {excel_filename}")

# Example usage
if __name__ == "__main__":
    # Sample data
    header = ['Employee', 'Department', 'Salary']
    data = [
        ['John Doe', 'Engineering', 75000],
        ['Jane Smith', 'Marketing', 65000],
        ['Mike Johnson', 'Sales', 70000]
    ]
    
    # Write Excel file
    write_excel_file('employees.xlsx', header, data)
    
    # Read Excel file
    excel_data = read_excel_file('employees.xlsx')
    
    # Convert CSV to Excel
    # First create a CSV file
    with open('temp.csv', 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(data)
    
    csv_to_excel('temp.csv', 'converted.xlsx')>>>>>>> External Changes
    
    
###################
### Directories ###
###################

# The following is for your information. The OS module will be discussed in COMP3006

import os

# https://docs.python.org/3/library/os.html

# Get current working directory

print(os.getcwd())

direct = os.getcwd() # we can also assign a directory to a variable

# change directory to a subfolder in current directory

os.chdir(os.path.join(direct, 'WK5_supplements')) # substitute the folder name with one of your own

print(os.getcwd()) # you should now be in the subdirectory folder

###############################
####------- Pathlib ------ ####
###############################

# The pathlib library provides a object oriented approach to directory navigation, etc. The syntax is more user friendly than the os package -in my opinion

# https://docs.python.org/3/library/pathlib.html

from pathlib import Path

# Get current working directory

current_dir = Path.cwd()
print(current_dir)

# Set cwd to one directory level up

# Change directory up one level
os.chdir(current_dir.parent) # for two levels use .parent.parent
print(current_dir)

'''
Some syntax with example below:

base_path = Path('path/to/directory') # absent raw text (r), use / or \\
filename = 'file.txt'
full_path = base_path / filename

'''

# Add a subdirectory to a base directory
base_dir = Path(r'C:\Users\Spring_4401\Lecture Notes and Materials') # I can use \ due to r, raw text
sub_dir = base_dir / 'WK5_supplements' # subdirectory in base_dir

# Change the current working directory to sub_dir
os.chdir(sub_dir)
print(Path.cwd())

# # To learn more about absolute and relative directory paths see:
# # https://towardsdatascience.com/simple-trick-to-work-with-relative-paths-in-python-c072cdc9acb9
# # https://towardsthecloud.com/get-relative-path-python

